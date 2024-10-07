import xlrd
import openpyxl

def read_excel_file(filename):
    """
    xls | xlsx | csv
    """
    if filename.endswith('xls'):
        workbook = xlrd.open_workbook(filename)
        sheet = workbook.sheet_by_index(0)
        # 列名
        column_names = [sheet.cell_value(0, col_idx) for col_idx in range(sheet.ncols)]
        for row_index in range(1, sheet.nrows):
            row_values = [sheet.cell_value(row_index, col_index) for col_index in range(sheet.ncols)]
            for column_name, value in zip(column_names, row_values):
                print(f"{column_name} : {value}")
            print()
    elif filename.endswith('xlsx'):
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.active
        # 列名
        column_names = [cell.value for cell in next(sheet.iter_rows())]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            for column_name, value in zip(column_names, row):
                print(f"{column_name}: {value}")
            print()
    elif filename.endswith('csv'):
        pass
    else:
        pass


if __name__ == "__main__":
    # filename = "static/upload/地址修复后模板.xls"
    # filename = 'static/upload/test.xlsx'
    # read_excel_file(filename)

    json = {
        '1': 1,
        '2': 2
    }

    json['1'] = 2
    print(json)
