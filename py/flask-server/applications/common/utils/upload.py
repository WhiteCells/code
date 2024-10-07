import os
import uuid

import pandas as pd
import requests
import openpyxl
import xlrd
import shutil
import json
from flask import current_app, jsonify, send_file, render_template, request
from flask_login import current_user
from sqlalchemy import desc
from applications.common.utils.http import fail_api, success_api
from applications.extensions import db
from applications.extensions.init_upload import photos
from applications.models import Photo, Excel
from applications.schemas import PhotoOutSchema, ExcelSchema
from applications.common.curd import model_to_dicts
from applications.config import payload, REPAIR_STATUS, ID_PREFIX


def get_photo(page, limit):
    photo = Photo.query.order_by(desc(Photo.create_time)).paginate(page=page, per_page=limit, error_out=False)
    count = Photo.query.count()
    data = model_to_dicts(schema=PhotoOutSchema, data=photo.items)
    return data, count


def upload_one(photo, mime):
    filename = photos.save(photo)
    file_url = '/_uploads/photos/' + filename
    upload_url = current_app.config.get("UPLOADED_PHOTOS_DEST")
    size = os.path.getsize(upload_url + '/' + filename)
    photo = Photo(name=filename, href=file_url, mime=mime, size=size)
    db.session.add(photo)
    db.session.commit()
    return file_url


def delete_photo_by_id(_id):
    photo_name = Photo.query.filter_by(id=_id).first().name
    photo = Photo.query.filter_by(id=_id).delete()
    db.session.commit()
    upload_url = current_app.config.get("UPLOADED_PHOTOS_DEST")
    os.remove(upload_url + '/' + photo_name)
    return photo


def get_excel(page, limit):
    excel = Excel.query.order_by(desc(Excel.create_time)).paginate(page=page, per_page=limit, error_out=False)
    count = Excel.query.count()
    data = model_to_dicts(schema=ExcelSchema, data=excel.items)
    return data, count


def upload_one_excel(excel, mime):
    filename = excel.filename
    file_path = os.path.join(current_app.config['UPLOADED_EXCELS_DEST'] + '/' + filename)

    excel.save(file_path)

    size = os.path.getsize(file_path)

    excel_entry = Excel(
        name=filename,
        href=file_path,
        mime=mime,
        size=size,
        repair_status='0',
        create_by=current_user.username,
        create_time=db.func.now()
    )
    db.session.add(excel_entry)
    db.session.commit()

    json = {
        'msg': '创建成功',
        'code': 0,
        'success': True,
        'data': {
            'src': file_path
        }
    }
    return jsonify(json)


def delete_excel_by_id(id):
    excel = Excel.query.filter_by(id=id).first()

    if not excel:
        return fail_api(msg='删除失败')

    upload_url = current_app.config.get("UPLOADED_EXCELS_DEST")
    file_path = os.path.join(upload_url, excel.name)

    try:
        db.session.delete(excel)
        db.session.commit()

        if os.path.exists(file_path):
            os.remove(file_path)

        return success_api(msg='删除成功')
    except Exception as e:
        db.session.rollback()
        print(e)
        return fail_api(msg='删除异常')


def download_excel_by_id(id):
    excel = Excel.query.filter_by(id=id).first()

    if not excel:
        return fail_api(msg='下载失败')

    file_name = excel.name
    # file_path = os.path.join(current_app.config.get('UPLOADED_EXCELS_DEST') + '/' + file_name)
    file_path = excel.href
    if not os.path.isfile(file_path):
        return fail_api(msg='文件未找到')

    try:
        return send_file(file_path, as_attachment=True)

    except Exception as e:
        print(e)
        return fail_api(msg='下载异常')


def read_excel2(file_path):
    """
    读取 CSV | XLS | XLSX 文件格式的文件
    :param file_path: 文件路径
    :return: 读取成功返回 DataFrame，失败返回 None
    """
    try:
        if file_path.endswith('.csv'):
            return pd.read_csv(file_path, dtype='str')
        elif file_path.endswith('.xls') or file_path.endswith('.xlsx'):
            return pd.read_excel(file_path, engine='openpyxl' if file_path.endswith('.xlsx') else None, dtype='str')
        else:
            return None
    except Exception as e:
        print(e)
        return None


def view_excel_by_id(id):
    excel = Excel.query.filter_by(id=id).first()
    try:
        file_url = excel.href

        df = read_excel2(file_url)

        if df is None:
            return fail_api('文件类型不支持或读取失败')

        html = df.to_html(index=False, classes='table table-striped')
        return render_template('system/excel/excel_view.html', df=html, filename=excel.name)
    except Exception as e:
        print(e)
        return fail_api('查看文件异常')


# def process_row(column_name, value, json_data):
#     if column_name == '客户':
#         json_data['entity_no'] = value
#     elif column_name == '委托方Id':
#         json_data['entrust_id'] = value
#     elif column_name == '任务Id':
#         json_data['task_id'] = value
#     elif column_name == '委案开始时间':
#         # base_date = datetime.datetime(1900, 1, 1)
#         # days = int(value)
#         # fraction = value - days
#         # time_seconds = fraction * 24 * 60 * 60
#         # time_delta = datetime.timedelta(seconds=time_seconds)
#         # result_date = base_date + datetime.timedelta(days=days - 1) + time_delta
#         #
#         # formatted_date = result_date.strftime('%Y-%m-%d %H:%M:%S')
#         # json_data['begin_time'] = formatted_date
#         json_data['begin_time'] = value.strftime('%Y/%m/%d %H:%M:%S')
#     elif column_name == '客户号':
#         json_data['customer_info']['customer_id'] = value
#     elif column_name == '家庭地址':
#         json_data['customer_info']['homeAddress'] = value
#     elif column_name == '委案开始时间':
#         json_data['customer_info']['beginTime'] = value
#     elif column_name == '居住地址':
#         json_data['customer_info']['liveAddress'] = value
#     elif column_name == '身份证':
#         json_data['customer_info']['user_identification'] = value
#     elif column_name == '其他地址1':
#         json_data['customer_info']['address1'] = value
#     elif column_name == '其他地址2':
#         json_data['customer_info']['address2'] = value


# def parse_excel_file(filename, excel):
#     """
#     xls | xlsx | csv
#     将读到的文件解析为 json
#     POST 发送给地址修复 API
#     回调时访问 receive 接口，将数据解析为 excel 保存至服务器
#     """
#
#     if filename.endswith('.xls'):
#         workbook = xlrd.open_workbook(filename)
#         sheet = workbook.sheet_by_index(0)
#         column_names = [sheet.cell_value(0, col_idx) for col_idx in range(sheet.ncols)]
#
#         for row_index in range(1, sheet.nrows):
#             json_data = {}
#             row_values = [sheet.cell_value(row_index, col_idx) for col_idx in range(sheet.ncols)]
#
#             for column_name, value in zip(column_names, row_values):
#                 process_row(column_name, value, json_data)
#
#             payload['returnUrl'] = current_app.config.get('RETURN_URL')
#             print(json_data)
#
#             # 调用 API，发送 JSON
#             response = requests.post(current_app.config.get('UPLOAD_URL'), json=json_data)
#             if response.status_code == '200' and response.json().get('code') == '200':
#                 excel.repair_status = REPAIR_STATUS['unrepaired']
#                 db.session.commit()
#             else:
#                 return fail_api(msg='调用修复 API 错误')
#         return success_api(msg='修复地址数据上传成功')
#
#     elif filename.endswith('.xlsx'):
#         workbook = openpyxl.load_workbook(filename)
#         sheet = workbook.active
#         column_names = [cell.value for cell in next(sheet.iter_rows())]
#
#         for row in sheet.iter_rows(min_row=2, values_only=True):
#             json_data = {}
#             for column_name, value in zip(column_names, row):
#                 process_row(column_name, value, json_data)
#
#             payload['returnUrl'] = current_app.config.get('RETURN_URL')
#             print(json_data)
#
#             # 调用 API，发送 JSON
#             response = requests.post(current_app.config.get('UPLOAD_URL'), json=json_data)
#             if response.status_code == '200' and response.json().get('code') == '200':
#                 excel.repair_status = REPAIR_STATUS['repairing']
#                 db.session.commit()
#             else:
#                 return fail_api(msg='调用修复 API 错误')
#         return success_api(msg='修复地址数据上传成功')
#
#     elif filename.endswith('.csv'):
#         return success_api(msg='csv todo')
#     else:
#         return fail_api(msg='修复地址数据上传失败')


def upload_excel_by_id(id):
    excel = Excel.query.filter_by(id=id).first()

    if not excel:
        return fail_api(msg='上传失败')

    file_name = excel.name
    file_path = excel.href
    id = excel.id

    id_str = ID_PREFIX + str(id)
    print('===> ', id_str)

    df = read_excel2(file_path)
    if df is None:
        return fail_api(msg='上传失败')

    # 对 dataframe 数据中的列名重命名
    df = df.rename(columns={
        '客户': 'entity_no',
        '委托方Id': 'entrust_id',
        '任务Id': 'task_id',
        '委案开始时间': 'begin_time',
        '客户号': 'customer_id',
        '家庭地址': 'homeAddress',
        '居住地址': 'liveAddress',
        '身份证': 'user_identification',
        '其他地址1': 'address1',
        '其他地址2': 'address2'
    })

    json_list = df.to_json(orient='records')
    print(json_list)
    parsed_json = json.loads(json_list)
    print(json.dumps(parsed_json, ensure_ascii=False, indent=2))

    uuid_str = str(uuid.uuid4())

    json_data = json.dumps({
        'entrust_id': uuid_str,
        'task_id': id_str,
        'returnUrl': current_app.config.get('RETURN_URL'),
        'customer_info': json_list
    })
    print(json_data)
    parsed_json = json.loads(json_data)
    print(json.dumps(parsed_json, ensure_ascii=False, indent=2))

    headers = {
        'Content-Type': 'application/json'
    }
    # 将 json 列表发送至 api
    response = requests.request('POST', current_app.config.get('REPAIR_API_URL'), headers=headers, data=json_data)
    if response.status_code == 200 and response.json().get('code') == '200':
        excel.repair_status = REPAIR_STATUS['repairing']
        db.session.commit()
        return success_api("修复请求已发送")

    return fail_api(msg='修复请求发送失败')


def api_receive_json():
    receive_json = request.get_json()
    # print(receive_json)
    id = receive_json['taskId']
    print(f'===> id: {id}')
    if id.startswith(ID_PREFIX):
        id = int(id[len(ID_PREFIX):])
    else:
        return fail_api(msg='ID 前缀匹配失败')

    repaired_address = receive_json.get('addressRepairs', [])

    if repaired_address:
        # 根据 id 查询文件名，file_name_id.xlsx
        excel = Excel.query.filter_by(id=id).first()
        if not excel:
            return fail_api(msg='文件索引不存在')

        file_name = excel.name
        file_without_extension_name = os.path.splitext(file_name)[0]
        print(file_without_extension_name)

        out_file_path = f'static/upload/result/{file_without_extension_name}_{id}.xlsx'
        print(out_file_path)

        address_repairs = receive_json['addressRepairs']

        df = pd.DataFrame(address_repairs)

        df = df.rename(columns={
            'entityNo': '客户',
            'entrustId': '委托方Id',
            'taskId': '任务Id',
            'custNo': '客户号',
            'registerAddress': '修复后的户籍地址',
            'liveAddress': '修复后的居住地址',
            'otherAddress1': '修复后的其他地址1',
            'otherAddress2': '修复后的其他地址2',
            'typeRegisterAddress': '地址修复模型返回的户籍地址类型',
            'typeLiveAddress': '地址修复模型返回的居住地址类型',
            'typeOtherAddress1': '地址修复模型返回的其他地址1类型',
            'typeOtherAddress2': '地址修复模型返回的其他地址2类型',
            'beginTime': '委案时间',
            'editAddress': '原户籍地址',
            'editAddress2': '原居住地址',
            'editAddress3': '原其他地址1',
            'editAddress4': '原其他地址2'
        })

        df.to_excel(out_file_path, index=False)

        # 修复完成后，修改数据库中的 excel 的索引，将其指向 result/xxx_id.xlsx
        excel.href = out_file_path
        excel.repair_status = REPAIR_STATUS['repaired']
        db.session.commit()

        return success_api(msg='修复成功')

    # 删除生成的模板
    return fail_api(msg='修复失败')


def download_template():
    file_path = current_app.config.get('TEMPLATE_EXCELS_DEST')
    if not os.path.isfile(file_path):
        return fail_api(msg='模板不存在')

    try:
        return send_file(file_path, as_attachment=True)

    except Exception as e:
        print(e)
        return fail_api(msg='下载失败')